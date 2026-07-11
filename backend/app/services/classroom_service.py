"""Class room management service."""
import io
from datetime import date

from fastapi import HTTPException, UploadFile
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload
from openpyxl import load_workbook

from app.models.student import ClassRoom, Student, Department
from app.models.user import User
from app.permissions.scoping import assert_department_access, get_allowed_department_ids
from app.repositories.audit_repository import AuditRepository
from app.schemas.classroom import ClassRoomCreate, ClassRoomUpdate, ClassRoomResponse, StudentImportResult


class ClassRoomService:
    def __init__(self, db: Session):
        self.db = db
        self.audit_repo = AuditRepository(db)

    def list_classes(self, user: User, department_id: int | None = None, cohort: str | None = None) -> list[ClassRoomResponse]:
        allowed = get_allowed_department_ids(user, self.db)
        query = self.db.query(ClassRoom).options(joinedload(ClassRoom.department))
        if department_id:
            if allowed is not None and department_id not in allowed:
                raise HTTPException(status_code=403, detail="Không có quyền xem chi đoàn này")
            query = query.filter(ClassRoom.department_id == department_id)
        elif allowed is not None:
            query = query.filter(ClassRoom.department_id.in_(allowed))
        if cohort:
            query = query.filter(ClassRoom.cohort == cohort)
        classes = query.order_by(ClassRoom.department_id, ClassRoom.name).all()
        return [self._to_response(c) for c in classes]

    def create(self, data: ClassRoomCreate, user: User) -> ClassRoomResponse:
        assert_department_access(user, self.db, data.department_id)
        existing = self.db.query(ClassRoom).filter(
            ClassRoom.department_id == data.department_id,
            ClassRoom.name == data.name,
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail="Lớp đã tồn tại trong chi đoàn này")
        cls = ClassRoom(**data.model_dump())
        self.db.add(cls)
        self.db.flush()
        self.audit_repo.log(user.id, "CREATE_CLASS", "class", cls.id, new_value={"name": data.name})
        self.db.commit()
        self.db.refresh(cls)
        return self._to_response(cls)

    def update(self, class_id: int, data: ClassRoomUpdate, user: User) -> ClassRoomResponse:
        cls = self._get_or_404(class_id)
        assert_department_access(user, self.db, cls.department_id)
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(cls, key, value)
        self.db.commit()
        self.db.refresh(cls)
        return self._to_response(cls)

    def delete(self, class_id: int, user: User):
        cls = self._get_or_404(class_id)
        assert_department_access(user, self.db, cls.department_id)
        count = self.db.query(func.count(Student.id)).filter(Student.class_id == class_id).scalar() or 0
        if count > 0:
            raise HTTPException(status_code=400, detail=f"Lớp còn {count} đoàn viên, không thể xóa")
        self.db.delete(cls)
        self.db.commit()

    async def import_students_excel(self, class_id: int, file: UploadFile, user: User, auto_create_books: bool = False) -> StudentImportResult:
        cls = self._get_or_404(class_id)
        assert_department_access(user, self.db, cls.department_id)
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        imported, skipped, errors = 0, 0, []
        from app.services.student_service import StudentService
        student_svc = StudentService(self.db)

        for idx, row in enumerate(rows, start=2):
            if not row or not row[0]:
                continue
            mssv = str(row[0]).strip()
            full_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            email = str(row[2]).strip() if len(row) > 2 and row[2] else None
            phone = str(row[3]).strip() if len(row) > 3 and row[3] else None
            if not full_name:
                errors.append(f"Dòng {idx}: thiếu họ tên")
                skipped += 1
                continue
            if self.db.query(Student).filter(Student.mssv == mssv).first():
                skipped += 1
                continue
            try:
                from app.schemas.student import StudentCreate
                student_svc.create(
                    StudentCreate(
                        mssv=mssv,
                        full_name=full_name,
                        cohort=cls.cohort,
                        class_id=cls.id,
                        department_id=cls.department_id,
                        email=email,
                        phone=phone,
                    ),
                    user.id,
                    skip_scope_check=True,
                )
                if auto_create_books:
                    from app.services.book_service import BookService
                    from app.schemas.book import BookCreate
                    s = self.db.query(Student).filter(Student.mssv == mssv).first()
                    if s and not s.union_book:
                        try:
                            BookService(self.db).create_book(BookCreate(student_id=s.id), user.id, skip_scope_check=True)
                        except HTTPException:
                            pass
                imported += 1
            except HTTPException as e:
                errors.append(f"Dòng {idx} ({mssv}): {e.detail}")
                skipped += 1
        return StudentImportResult(imported=imported, skipped=skipped, errors=errors[:20])

    def _get_or_404(self, class_id: int) -> ClassRoom:
        cls = self.db.query(ClassRoom).options(joinedload(ClassRoom.department)).filter(ClassRoom.id == class_id).first()
        if not cls:
            raise HTTPException(status_code=404, detail="Lớp không tồn tại")
        return cls

    def _to_response(self, cls: ClassRoom) -> ClassRoomResponse:
        count = self.db.query(func.count(Student.id)).filter(Student.class_id == cls.id).scalar() or 0
        return ClassRoomResponse(
            id=cls.id,
            name=cls.name,
            cohort=cls.cohort,
            department_id=cls.department_id,
            department_name=cls.department.name if cls.department else None,
            student_count=count,
        )
