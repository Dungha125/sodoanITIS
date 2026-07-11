"""Student management service."""
import io
from datetime import date

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.student import Student, ClassRoom, Department
from app.models.user import User
from app.repositories.student_repository import StudentRepository
from app.repositories.audit_repository import AuditRepository
from app.permissions.scoping import assert_department_access, assert_student_access, resolve_department_filter
from app.permissions.roles import ROLE_BI_THU, ROLE_LIEN_CHI_DOAN, ROLE_SUPER_ADMIN
from app.schemas.student import StudentCreate, StudentUpdate, StudentResponse, StudentStatusUpdate
from app.services.period_service import CollectionPeriodService

GENDER_MAP = {"nam": "Nam", "nữ": "Nữ", "nu": "Nữ", "male": "Nam", "female": "Nữ", "m": "Nam", "f": "Nữ"}


class StudentService:
    def __init__(self, db: Session):
        self.db = db
        self.repo = StudentRepository(db)
        self.audit_repo = AuditRepository(db)

    def create(self, data: StudentCreate, actor_id: int, user: User | None = None, skip_scope_check: bool = False) -> StudentResponse:
        if not skip_scope_check and user:
            assert_department_access(user, self.db, data.department_id)
        if self.repo.get_by_mssv(data.mssv):
            raise HTTPException(status_code=409, detail="MSSV đã tồn tại")
        self._validate_class_department(data.class_id, data.department_id)
        payload = data.model_dump()
        payload["cohort"] = payload.get("cohort") or self._cohort_name(data.department_id)
        payload["gender"] = self._normalize_gender(payload.get("gender"))
        student = Student(**payload)
        self.repo.create(student)
        self.audit_repo.log(actor_id, "CREATE_STUDENT", "student", student.id, new_value={"mssv": data.mssv})
        self.repo.commit()
        student = self.repo.get_by_id(student.id)
        return self._to_response(student)

    def update(self, student_id: int, data: StudentUpdate, actor_id: int, user: User) -> StudentResponse:
        student = self.repo.get_by_id(student_id)
        if not student:
            raise HTTPException(status_code=404, detail="Đoàn viên không tồn tại")
        assert_student_access(user, self.db, student)
        if data.department_id is not None:
            assert_department_access(user, self.db, data.department_id)
        new_dept = data.department_id if data.department_id is not None else student.department_id
        new_class = data.class_id if data.class_id is not None else student.class_id
        if data.class_id is not None or data.department_id is not None:
            self._validate_class_department(new_class, new_dept)
        old = {"mssv": student.mssv, "full_name": student.full_name}
        updates = data.model_dump(exclude_unset=True)
        if "gender" in updates:
            updates["gender"] = self._normalize_gender(updates["gender"])
        for key, value in updates.items():
            setattr(student, key, value)
        if data.department_id and not data.cohort:
            student.cohort = self._cohort_name(student.department_id)
        self.audit_repo.log(actor_id, "UPDATE_STUDENT", "student", student.id, old_value=old, new_value=updates)
        self.repo.commit()
        return self._to_response(student)

    def update_status(self, student_id: int, data: StudentStatusUpdate, actor_id: int, user: User) -> StudentResponse:
        student = self.repo.get_by_id(student_id)
        if not student:
            raise HTTPException(status_code=404, detail="Đoàn viên không tồn tại")
        assert_student_access(user, self.db, student)

        if user.role.code == ROLE_BI_THU:
            if not CollectionPeriodService(self.db).is_update_open():
                raise HTTPException(status_code=403, detail="Ngoài khoảng thời gian cập nhật trạng thái")
            if data.book_submitted is None and data.fee_submitted is None:
                raise HTTPException(status_code=400, detail="Cần chọn trạng thái cập nhật")

        old = {"book_submitted": student.book_submitted, "fee_submitted": student.fee_submitted}
        if data.book_submitted is not None:
            student.book_submitted = data.book_submitted
        if data.fee_submitted is not None:
            student.fee_submitted = data.fee_submitted
        self.audit_repo.log(actor_id, "UPDATE_STUDENT_STATUS", "student", student.id, old_value=old,
                            new_value={"book_submitted": student.book_submitted, "fee_submitted": student.fee_submitted})
        self.repo.commit()
        return self._to_response(student)

    def delete(self, student_id: int, actor_id: int, user: User):
        student = self.repo.get_by_id(student_id)
        if not student:
            raise HTTPException(status_code=404, detail="Đoàn viên không tồn tại")
        assert_student_access(user, self.db, student)
        self.audit_repo.log(actor_id, "DELETE_STUDENT", "student", student.id)
        self.repo.delete(student)
        self.repo.commit()

    def get(self, student_id: int, user: User) -> StudentResponse:
        student = self.repo.get_by_id(student_id)
        if not student:
            raise HTTPException(status_code=404, detail="Đoàn viên không tồn tại")
        assert_student_access(user, self.db, student)
        return self._to_response(student)

    def list(self, user: User, page=1, size=20, **filters) -> dict:
        dept_ids = resolve_department_filter(user, self.db, filters.pop("department_id", None))
        if dept_ids is not None and len(dept_ids) == 0:
            return {"items": [], "total": 0, "page": page, "size": size, "pages": 0}
        skip = (page - 1) * size
        items, total = self.repo.search(skip=skip, limit=size, department_ids=dept_ids, **filters)
        return {
            "items": [self._to_response(s) for s in items],
            "total": total,
            "page": page,
            "size": size,
            "pages": (total + size - 1) // size if size else 0,
        }

    async def import_excel(self, department_id: int, file: UploadFile, user: User) -> dict:
        assert_department_access(user, self.db, department_id)
        dept = self.db.query(Department).filter(Department.id == department_id).first()
        if not dept:
            raise HTTPException(status_code=404, detail="Chi đoàn không tồn tại")

        content = await file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        imported, skipped, errors = 0, 0, []
        cohort_name = self._cohort_name(department_id)

        for idx, row in enumerate(rows, start=2):
            if not row or not row[0]:
                continue
            mssv = str(row[0]).strip()
            full_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            dob_raw = row[2] if len(row) > 2 else None
            gender_raw = str(row[3]).strip() if len(row) > 3 and row[3] else None
            phone = str(row[4]).strip() if len(row) > 4 and row[4] else None

            if not full_name:
                errors.append(f"Dòng {idx}: thiếu họ tên")
                skipped += 1
                continue
            if self.repo.get_by_mssv(mssv):
                skipped += 1
                continue

            dob = None
            if isinstance(dob_raw, date):
                dob = dob_raw
            elif dob_raw:
                try:
                    dob = date.fromisoformat(str(dob_raw)[:10])
                except ValueError:
                    pass

            try:
                self.create(
                    StudentCreate(
                        mssv=mssv, full_name=full_name, date_of_birth=dob,
                        gender=gender_raw, phone=phone, department_id=department_id,
                        cohort=cohort_name,
                    ),
                    user.id, user, skip_scope_check=True,
                )
                imported += 1
            except HTTPException as e:
                errors.append(f"Dòng {idx}: {e.detail}")
                skipped += 1

        return {"imported": imported, "skipped": skipped, "errors": errors[:20]}

    def _validate_class_department(self, class_id: int | None, department_id: int):
        if not class_id:
            return
        cls = self.db.query(ClassRoom).filter(ClassRoom.id == class_id).first()
        if not cls:
            raise HTTPException(status_code=404, detail="Lớp không tồn tại")
        if cls.department_id != department_id:
            raise HTTPException(status_code=400, detail="Lớp không thuộc chi đoàn đã chọn")

    def _cohort_name(self, department_id: int) -> str:
        from sqlalchemy.orm import joinedload
        dept = self.db.query(Department).options(joinedload(Department.cohort)).filter(Department.id == department_id).first()
        if dept and dept.cohort:
            return dept.cohort.name
        return "UNKNOWN"

    @staticmethod
    def _normalize_gender(value: str | None) -> str | None:
        if not value:
            return None
        return GENDER_MAP.get(value.strip().lower(), value.strip())

    def _to_response(self, student: Student) -> StudentResponse:
        if student.department is None and student.department_id:
            self.db.refresh(student, ["department", "class_room"])
        return StudentResponse(
            id=student.id,
            mssv=student.mssv,
            full_name=student.full_name,
            date_of_birth=student.date_of_birth,
            gender=student.gender,
            cohort=student.cohort,
            class_id=student.class_id,
            department_id=student.department_id,
            union_join_date=student.union_join_date,
            admission_date=student.admission_date,
            email=student.email,
            phone=student.phone,
            book_submitted=student.book_submitted,
            fee_submitted=student.fee_submitted,
            status=student.status,
            created_at=student.created_at,
            department_name=student.department.name if student.department else None,
            class_name=student.class_room.name if student.class_room else None,
            book_status_label="Đã nộp" if student.book_submitted else "Chưa nộp",
            fee_status_label="Đã nộp" if student.fee_submitted else "Chưa nộp",
            gender_label=student.gender,
        )
